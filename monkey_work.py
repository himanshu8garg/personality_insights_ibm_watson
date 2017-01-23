import json
import math
#have to install openpyxl package using pip
import openpyxl
#have to install requests package using pip
import requests
#have to pip install --upgrade watson-developer-cloud
from watson_developer_cloud import PersonalityInsightsV3

#--------------------------------------------------------------
#using Watson service to generate a json object which is to be used later to fill in the spreadsheet
personality_insights = PersonalityInsightsV3(
   version='2016-12-15',
   username='ed46916e-0566-45d6-8f8b-830d60fc6a42',
   password='D13MUx1G1tLF',
   #x-watson-learning-opt-out=True
 )
with open('./profile.json') as profile_json:
   profile = personality_insights.profile(
     profile_json.read(), content_type='application/json',
     raw_scores=True, consumption_preferences=True)
  
print(json.dumps(profile, indent=2))
#----------------------------------------------------------------------------------------------
json_data=json.dumps(json.loads("""{
  "word_count": 2653,
  "processed_language": "en",
  "personality": [
    {
      "trait_id": "big5_openness",
      "name": "Openness",
      "category": "personality",
      "percentile": 0.968260901812302,
      "children": [
        {
          "trait_id": "facet_adventurousness",
          "name": "Adventurousness",
          "category": "personality",
          "percentile": 0.9015378969727104
        },
        {
          "trait_id": "facet_artistic_interests",
          "name": "Artistic interests",
          "category": "personality",
          "percentile": 0.5950633257241809
        },
        {
          "trait_id": "facet_emotionality",
          "name": "Emotionality",
          "category": "personality",
          "percentile": 0.14633483847211298
        },
        {
          "trait_id": "facet_imagination",
          "name": "Imagination",
          "category": "personality",
          "percentile": 0.02479252007794369
        },
        {
          "trait_id": "facet_intellect",
          "name": "Intellect",
          "category": "personality",
          "percentile": 0.9752863601907291
        },
        {
          "trait_id": "facet_liberalism",
          "name": "Authority-challenging",
          "category": "personality",
          "percentile": 0.9293826299227845
        }
      ]
    },
    {
      "trait_id": "big5_conscientiousness",
      "name": "Conscientiousness",
      "category": "personality",
      "percentile": 0.8690018318252943,
      "children": [
        {
          "trait_id": "facet_achievement_striving",
          "name": "Achievement striving",
          "category": "personality",
          "percentile": 0.8482229929630647
        },
        {
          "trait_id": "facet_cautiousness",
          "name": "Cautiousness",
          "category": "personality",
          "percentile": 0.9879375379001101
        },
        {
          "trait_id": "facet_dutifulness",
          "name": "Dutifulness",
          "category": "personality",
          "percentile": 0.9113684347779039
        },
        {
          "trait_id": "facet_orderliness",
          "name": "Orderliness",
          "category": "personality",
          "percentile": 0.3402774902580711
        },
        {
          "trait_id": "facet_self_discipline",
          "name": "Self-discipline",
          "category": "personality",
          "percentile": 0.8676817059043762
        },
        {
          "trait_id": "facet_self_efficacy",
          "name": "Self-efficacy",
          "category": "personality",
          "percentile": 0.6636347981531293
        }
      ]
    },
    {
      "trait_id": "big5_extraversion",
      "name": "Extraversion",
      "category": "personality",
      "percentile": 0.8398052833471685,
      "children": [
        {
          "trait_id": "facet_activity_level",
          "name": "Activity level",
          "category": "personality",
          "percentile": 0.9809212147463013
        },
        {
          "trait_id": "facet_assertiveness",
          "name": "Assertiveness",
          "category": "personality",
          "percentile": 0.9610846443932104
        },
        {
          "trait_id": "facet_cheerfulness",
          "name": "Cheerfulness",
          "category": "personality",
          "percentile": 0.39506222832205873
        },
        {
          "trait_id": "facet_excitement_seeking",
          "name": "Excitement-seeking",
          "category": "personality",
          "percentile": 0.01557678027182019
        },
        {
          "trait_id": "facet_friendliness",
          "name": "Outgoing",
          "category": "personality",
          "percentile": 0.8878734834544777
        },
        {
          "trait_id": "facet_gregariousness",
          "name": "Gregariousness",
          "category": "personality",
          "percentile": 0.33433436993784066
        }
      ]
    },
    {
      "trait_id": "big5_agreeableness",
      "name": "Agreeableness",
      "category": "personality",
      "percentile": 0.1886942344971127,
      "children": [
        {
          "trait_id": "facet_altruism",
          "name": "Altruism",
          "category": "personality",
          "percentile": 0.8467772968967862
        },
        {
          "trait_id": "facet_cooperation",
          "name": "Cooperation",
          "category": "personality",
          "percentile": 0.9435898467131107
        },
        {
          "trait_id": "facet_modesty",
          "name": "Modesty",
          "category": "personality",
          "percentile": 0.3375063991225501
        },
        {
          "trait_id": "facet_morality",
          "name": "Uncompromising",
          "category": "personality",
          "percentile": 0.6728155118276474
        },
        {
          "trait_id": "facet_sympathy",
          "name": "Sympathy",
          "category": "personality",
          "percentile": 0.9587426183239258
        },
        {
          "trait_id": "facet_trust",
          "name": "Trust",
          "category": "personality",
          "percentile": 0.9669047607843633
        }
      ]
    },
    {
      "trait_id": "big5_neuroticism",
      "name": "Emotional range",
      "category": "personality",
      "percentile": 0.9366368389828348,
      "children": [
        {
          "trait_id": "facet_anger",
          "name": "Fiery",
          "category": "personality",
          "percentile": 0.012436958580311186
        },
        {
          "trait_id": "facet_anxiety",
          "name": "Prone to worry",
          "category": "personality",
          "percentile": 0.029079553995418783
        },
        {
          "trait_id": "facet_depression",
          "name": "Melancholy",
          "category": "personality",
          "percentile": 0.1491848820989652
        },
        {
          "trait_id": "facet_immoderation",
          "name": "Immoderation",
          "category": "personality",
          "percentile": 0.031421556471662226
        },
        {
          "trait_id": "facet_self_consciousness",
          "name": "Self-consciousness",
          "category": "personality",
          "percentile": 0.07500495809941077
        },
        {
          "trait_id": "facet_vulnerability",
          "name": "Susceptible to stress",
          "category": "personality",
          "percentile": 0.06802442307086404
        }
      ]
    }
  ],
  "needs": [
    {
      "trait_id": "need_challenge",
      "name": "Challenge",
      "category": "needs",
      "percentile": 0.055189818208056685
    },
    {
      "trait_id": "need_closeness",
      "name": "Closeness",
      "category": "needs",
      "percentile": 0.11650737857203425
    },
    {
      "trait_id": "need_curiosity",
      "name": "Curiosity",
      "category": "needs",
      "percentile": 0.2125786321441694
    },
    {
      "trait_id": "need_excitement",
      "name": "Excitement",
      "category": "needs",
      "percentile": 0.02989232795759339
    },
    {
      "trait_id": "need_harmony",
      "name": "Harmony",
      "category": "needs",
      "percentile": 0.030309378629459705
    },
    {
      "trait_id": "need_ideal",
      "name": "Ideal",
      "category": "needs",
      "percentile": 0.028415562625504376
    },
    {
      "trait_id": "need_liberty",
      "name": "Liberty",
      "category": "needs",
      "percentile": 0.0386417136753055
    },
    {
      "trait_id": "need_love",
      "name": "Love",
      "category": "needs",
      "percentile": 0.02846041532512933
    },
    {
      "trait_id": "need_practicality",
      "name": "Practicality",
      "category": "needs",
      "percentile": 0.08058749554433037
    },
    {
      "trait_id": "need_self_expression",
      "name": "Self-expression",
      "category": "needs",
      "percentile": 0.005056740661799508
    },
    {
      "trait_id": "need_stability",
      "name": "Stability",
      "category": "needs",
      "percentile": 0.04787811949694171
    },
    {
      "trait_id": "need_structure",
      "name": "Structure",
      "category": "needs",
      "percentile": 0.3979258686579211
    }
  ],
  "values": [
    {
      "trait_id": "value_conservation",
      "name": "Conservation",
      "category": "values",
      "percentile": 0.02568967674997019
    },
    {
      "trait_id": "value_openness_to_change",
      "name": "Openness to change",
      "category": "values",
      "percentile": 0.36059207299374263
    },
    {
      "trait_id": "value_hedonism",
      "name": "Hedonism",
      "category": "values",
      "percentile": 0.010684744091107035
    },
    {
      "trait_id": "value_self_enhancement",
      "name": "Self-enhancement",
      "category": "values",
      "percentile": 0.04251309238202783
    },
    {
      "trait_id": "value_self_transcendence",
      "name": "Self-transcendence",
      "category": "values",
      "percentile": 0.03317450832422442
    }
  ],
  "consumption_preferences": [
    {
      "consumption_preference_category_id": "consumption_preferences_shopping",
      "name": "Purchasing Preferences",
      "consumption_preferences": [
        {
          "consumption_preference_id": "consumption_preferences_automobile_ownership_cost",
          "name": "Likely to be sensitive to ownership cost when buying automobiles",
          "score": 1
        },
        {
          "consumption_preference_id": "consumption_preferences_automobile_safety",
          "name": "Likely to prefer safety when buying automobiles",
          "score": 1
        },
        {
          "consumption_preference_id": "consumption_preferences_automobile_resale_value",
          "name": "Likely to prefer resale value when buying automobiles",
          "score": 1
        },
        {
          "consumption_preference_id": "consumption_preferences_clothes_quality",
          "name": "Likely to prefer quality when buying clothes",
          "score": 1
        },
        {
          "consumption_preference_id": "consumption_preferences_clothes_style",
          "name": "Likely to prefer style when buying clothes",
          "score": 0
        },
        {
          "consumption_preference_id": "consumption_preferences_clothes_comfort",
          "name": "Likely to prefer comfort when buying clothes",
          "score": 1
        },
        {
          "consumption_preference_id": "consumption_preferences_influence_brand_name",
          "name": "Likely to be influenced by brand name when making product purchases",
          "score": 0
        },
        {
          "consumption_preference_id": "consumption_preferences_influence_utility",
          "name": "Likely to be influenced by product utility when making product purchases",
          "score": 0
        },
        {
          "consumption_preference_id": "consumption_preferences_influence_online_ads",
          "name": "Likely to be influenced by online ads when making product purchases",
          "score": 0
        },
        {
          "consumption_preference_id": "consumption_preferences_influence_social_media",
          "name": "Likely to be influenced by social media when making product purchases",
          "score": 0
        },
        {
          "consumption_preference_id": "consumption_preferences_influence_family_members",
          "name": "Likely to be influenced by family when making product purchases",
          "score": 1
        },
        {
          "consumption_preference_id": "consumption_preferences_spur_of_moment",
          "name": "Likely to indulge in spur of the moment purchases",
          "score": 0
        },
        {
          "consumption_preference_id": "consumption_preferences_credit_card_payment",
          "name": "Likely to prefer using credit cards for shopping",
          "score": 1
        }
      ]
    },
    {
      "consumption_preference_category_id": "consumption_preferences_health_and_activity",
      "name": "Health & Activity Preferences",
      "consumption_preferences": [
        {
          "consumption_preference_id": "consumption_preferences_eat_out",
          "name": "Likely to eat out frequently",
          "score": 0
        },
        {
          "consumption_preference_id": "consumption_preferences_fast_food_frequency",
          "name": "Likely to eat fast food frequently",
          "score": 0
        },
        {
          "consumption_preference_id": "consumption_preferences_gym_membership",
          "name": "Likely to have a gym membership",
          "score": 0
        },
        {
          "consumption_preference_id": "consumption_preferences_adventurous_sports",
          "name": "Likely to like adventurous sports",
          "score": 0
        },
        {
          "consumption_preference_id": "consumption_preferences_outdoor",
          "name": "Likely to like outdoor activities",
          "score": 1
        }
      ]
    },
    {
      "consumption_preference_category_id": "consumption_preferences_environmental_concern",
      "name": "Environmental Concern Preferences",
      "consumption_preferences": [
        {
          "consumption_preference_id": "consumption_preferences_concerned_environment",
          "name": "Likely to be concerned about the environment",
          "score": 1
        }
      ]
    },
    {
      "consumption_preference_category_id": "consumption_preferences_entrepreneurship",
      "name": "Entrepreneurship Preferences",
      "consumption_preferences": [
        {
          "consumption_preference_id": "consumption_preferences_start_business",
          "name": "Likely to consider starting a business in next few years",
          "score": 0.5
        }
      ]
    },
    {
      "consumption_preference_category_id": "consumption_preferences_movie",
      "name": "Movie Preferences",
      "consumption_preferences": [
        {
          "consumption_preference_id": "consumption_preferences_movie_romance",
          "name": "Likely to like romance movies",
          "score": 0
        },
        {
          "consumption_preference_id": "consumption_preferences_movie_adventure",
          "name": "Likely to like adventure movies",
          "score": 1
        },
        {
          "consumption_preference_id": "consumption_preferences_movie_horror",
          "name": "Likely to like horror movies",
          "score": 0
        },
        {
          "consumption_preference_id": "consumption_preferences_movie_musical",
          "name": "Likely to like musical movies",
          "score": 0
        },
        {
          "consumption_preference_id": "consumption_preferences_movie_historical",
          "name": "Likely to like historical movies",
          "score": 1
        },
        {
          "consumption_preference_id": "consumption_preferences_movie_science_fiction",
          "name": "Likely to like science-fiction movies",
          "score": 1
        },
        {
          "consumption_preference_id": "consumption_preferences_movie_war",
          "name": "Likely to like war movies",
          "score": 1
        },
        {
          "consumption_preference_id": "consumption_preferences_movie_drama",
          "name": "Likely to like drama movies",
          "score": 0
        },
        {
          "consumption_preference_id": "consumption_preferences_movie_action",
          "name": "Likely to like action movies",
          "score": 1
        },
        {
          "consumption_preference_id": "consumption_preferences_movie_documentary",
          "name": "Likely to like documentary movies",
          "score": 1
        }
      ]
    },
    {
      "consumption_preference_category_id": "consumption_preferences_music",
      "name": "Music Preferences",
      "consumption_preferences": [
        {
          "consumption_preference_id": "consumption_preferences_music_rap",
          "name": "Likely to like rap music",
          "score": 0
        },
        {
          "consumption_preference_id": "consumption_preferences_music_country",
          "name": "Likely to like country music",
          "score": 0.5
        },
        {
          "consumption_preference_id": "consumption_preferences_music_r_b",
          "name": "Likely to like R&B music",
          "score": 0.5
        },
        {
          "consumption_preference_id": "consumption_preferences_music_hip_hop",
          "name": "Likely to like hip hop music",
          "score": 0
        },
        {
          "consumption_preference_id": "consumption_preferences_music_live_event",
          "name": "Likely to attend live musical events",
          "score": 1
        },
        {
          "consumption_preference_id": "consumption_preferences_music_playing",
          "name": "Likely to have experience playing music",
          "score": 0
        },
        {
          "consumption_preference_id": "consumption_preferences_music_latin",
          "name": "Likely to like Latin music",
          "score": 1
        },
        {
          "consumption_preference_id": "consumption_preferences_music_rock",
          "name": "Likely to like rock music",
          "score": 1
        },
        {
          "consumption_preference_id": "consumption_preferences_music_classical",
          "name": "Likely to like classical music",
          "score": 1
        }
      ]
    },
    {
      "consumption_preference_category_id": "consumption_preferences_reading",
      "name": "Reading Preferences",
      "consumption_preferences": [
        {
          "consumption_preference_id": "consumption_preferences_read_frequency",
          "name": "Likely to read often",
          "score": 1
        },
        {
          "consumption_preference_id": "consumption_preferences_read_motive_enjoyment",
          "name": "Likely to read for enjoyment",
          "score": 1
        },
        {
          "consumption_preference_id": "consumption_preferences_read_motive_information",
          "name": "Likely to read for information",
          "score": 1
        },
        {
          "consumption_preference_id": "consumption_preferences_books_entertainment_magazines",
          "name": "Likely to read entertainment magazines",
          "score": 0
        },
        {
          "consumption_preference_id": "consumption_preferences_books_non_fiction",
          "name": "Likely to read non-fiction books",
          "score": 1
        },
        {
          "consumption_preference_id": "consumption_preferences_read_motive_mandatory",
          "name": "Likely to do mandatory reading only",
          "score": 0
        },
        {
          "consumption_preference_id": "consumption_preferences_read_motive_relaxation",
          "name": "Likely to read for relaxation",
          "score": 1
        },
        {
          "consumption_preference_id": "consumption_preferences_books_financial_investing",
          "name": "Likely to read financial investment books",
          "score": 1
        },
        {
          "consumption_preference_id": "consumption_preferences_books_autobiographies",
          "name": "Likely to read autobiographical books",
          "score": 1
        }
      ]
    },
    {
      "consumption_preference_category_id": "consumption_preferences_volunteering",
      "name": "Volunteering Preferences",
      "consumption_preferences": [
        {
          "consumption_preference_id": "consumption_preferences_volunteer",
          "name": "Likely to volunteer for social causes",
          "score": 1
        },
        {
          "consumption_preference_id": "consumption_preferences_volunteering_time",
          "name": "Likely to have spent time volunteering",
          "score": 1
        },
        {
          "consumption_preference_id": "consumption_preferences_volunteer_learning",
          "name": "Likely to volunteer to learn about social causes",
          "score": 1
        }
      ]
    }
  ],
  "warnings": []
}"""))

parsed_json=json.loads(json_data)
personality=[]
percentile=[]
for eachpersonality in parsed_json['personality']:
	personality.append(eachpersonality['name'])
	percentile.append(round((eachpersonality['percentile']*100),0))
	for eachsubpersonality in eachpersonality['children']:
		personality.append(eachsubpersonality['name'])
		percentile.append(round((eachsubpersonality['percentile']*100),0))
for eachpersonality in parsed_json['needs']:
	personality.append(eachpersonality['name'])
	percentile.append(round((eachpersonality['percentile']*100),0))
for eachpersonality in parsed_json['values']:
		personality.append(eachpersonality['name'])
		percentile.append(round((eachpersonality['percentile']*100),0))

from openpyxl import load_workbook
wb=load_workbook(filename='Watson-data.xlsx')
ws=wb.worksheets[0]
row_x=int(raw_input("Ina, please enter the row in which you want the data to magically appear: "))
ws["I"+str(row_x)]=parsed_json['word_count']
for row in ws['K1:BJ1']:
	for cell in row:
		for i in range(0,len(personality)):
			if cell.value==personality[i]:
				index=str(cell.column)+str(row_x)
				#print index
				print personality[i]
				print percentile[i]
				print cell.value
				print "-----------------------"
				ws[index]=percentile[i]
				#print cell.column
				#print percentile[i]

for i in range (0,len(personality)):
	print personality[i]
for row in ws['K1:BJ1']:
	for cell in row:
		if cell.value not in personality:
			print "----------------------------------------"
			print cell.value
			index=str(cell.column)+str(row_x)
			ws[index]=-1


wb.save(filename='Watson-data.xlsx')

